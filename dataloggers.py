#!python3
# dataloggers.py - format kestrel data

import csv
import json
import os
import shutil
import openpyxl
import requests
from time import mktime, strptime


class Kestrel:

    def __init__(self, host):
        self.host = host

    def mk_workbook(self, excel_file):
        try:
            self.wb = openpyxl.load_workbook(excel_file)
            return self.wb
        except:
            print('Provide valid kestrel excel sheet')
            return None

    def get_sheetnames(self, wb):
        try:
            self.sheetnames = wb.sheetnames
            return self.sheetnames
        except:
            print('provide excel workbook')
            return None

    def validate_sheet(self, wb, sheet):
        if wb[sheet].cell(row=1, column=1).value.lower().replace(' ', '') == 'devicename':
            if wb[sheet].cell(row=2, column=1).value.lower().replace(' ', '') == 'devicemodel':
                if wb[sheet].cell(row=3, column=1).value.lower().replace(' ', '') == 'serialnumber':
                    return True
        return False

    def get_data(self, wb, sheet):
        if self.validate_sheet(wb, sheet):
            formatted = {
                'host': self.host,
                'data_name': f'{self.host}_history',
                'data': []
            }
            for row in range(6, wb[sheet].max_row+1):
                created_at = self.get_epoch(wb[sheet].cell(row=row, column=1).value)
                entry = {
                    'created_at': created_at,
                    'data': {}
                }
                for column in range(2, wb[sheet].max_column+1):
                    key = wb[sheet].cell(row=4, column=column).value.replace(' ', '').lower()
                    value = float(wb[sheet].cell(row=row, column=column).value)
                    entry['data'][key] = value
                formatted['data'].append(entry)
            return formatted
        else:
            return None

    @staticmethod
    def get_epoch(time_val):
        # convert kestrel time val to epoch timestamp
        return int(time_val.timestamp())


class PurpleAir:

    def __init__(self, host, purple_id):
        self.host = host
        self.purple_id = purple_id
        self.live_page = f'https://www.purpleair.com/json?show={self.purple_id}'

    def __repr__(self):
        return f'PurpleAir("{self.host}", "{self.purple_id}")'

    def fetch_raw(self):
        # retrieve raw live data dictionary
        purple_response = requests.get(self.live_page)
        raw_data = json.loads(purple_response.content)
        raw_data['Stats'] = json.loads(raw_data['results'][0]['Stats'])

        return raw_data

    def fetch_data(self):
        # return formatted data dictionary
        raw_data = self.fetch_raw()
        formatted = []
        
        for entry in raw_data['results']:
            formatted.append(
                {
                    'purple_host': self.host,
                    'purple_id': entry.get('ID'),
                    'purple_name': entry.get('Label'),
                    'parent_id': entry.get('ParentID'),
                    'created_at': entry.get('LastSeen'),
                    'data': {
                        'rssi': float(entry.get('RSSI')),
                        'adc': float(entry.get('Adc')),
                        'p_0_3_um': float(entry.get('p_0_3_um')),
                        'p_0_5_um': float(entry.get('p_0_5_um')),
                        'p_1_0_um': float(entry.get('p_1_0_um')),
                        'p_2_5_um': float(entry.get('p_2_5_um')),
                        'p_5_0_um': float(entry.get('p_5_0_um')),
                        'p_10_0_um': float(entry.get('p_10_0_um')),
                        'pm1_0_cf_1': float(entry.get('pm1_0_cf_1')),
                        'pm2_5_cf_1': float(entry.get('pm2_5_cf_1')),
                        'pm10_0_cf_1': float(entry.get('pm10_0_cf_1')),
                        'pm1_0_atm': float(entry.get('pm1_0_atm')),
                        'pm2_5_atm': float(entry.get('pm2_5_atm')),
                        'pm10_0_atm': float(entry.get('pm10_0_atm')),
                        'humidity': int(entry.get('humidity')),
                        'temp_f': float(entry.get('temp_f')),
                        'pressure': float(entry.get('pressure')),
                    }
                }
            )
        
        return formatted

    def get_data(self, csv_file):
        # return list of formatted dictionaries
        meta, data = self.load_csv(csv_file)
        if data:
            try:
                formatted = {
                    'host': self.host,
                    'data_name': f'{self.host}_history',
                    'data': []
                }
                for line in data:
                    entry = {
                        'created_at': self.get_epoch(line[0]),
                        'data': {}
                    }
                    for i, column in enumerate(meta[1:]):
                        column = column.replace('.', '_')\
                            .replace('>=', 'p_')\
                            .replace('_ug/m3', '')\
                            .replace('_hpa', '')\
                            .replace('um/dl', '')\
                            .lower()
                        try:
                            value = int(line[i+1])
                        except ValueError:
                            try:
                                value = float(line[i+1])
                            except ValueError:
                                value = line[i+1]
                        entry['data'][column] = value
                    formatted['data'].append(entry)
                return formatted
            except:
                print('CSV format error!')
                return None
        else:
            print('Could not retrieve data from csv!')
            return None

    @staticmethod
    def load_csv(csv_file):
        try:
            with open(csv_file, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)
                meta, data = data[0], data[1:]
                meta.remove('')
            return (meta, data)
        except:
            return None

    @staticmethod
    def get_epoch(time_val):
        # convert purple time elem to epoch timestamp
        time_val = time_val.strip(' UTC')
        pattern = '%Y-%m-%d %H:%M:%S'
        epoch = int(mktime(strptime(time_val, pattern)))

        # convert epoch
        epoch -= 18_000
        return epoch


class CampbellSci:

    def __init__(self, host):
        self.host = host

    def fetch_raw(self, ip, table):
        page = f'http://{ip}/?command=dataquery&uri=dl:{table}&format=json&mode=most-recent'
        response = requests.get(page)
        raw_data = json.loads(response)

        return raw_data


    def get_data(self, dat_file):
        # return list of formatted dictionaries
        meta, data = self.load_csv(dat_file)
        if data:
            try:
                formatted = {
                    'host': self.host,
                    'data_name': f'{self.host}_history',
                    'data': []
                }
                for line in data:
                    entry = {
                        'created_at': self.get_epoch(line[0]),
                        'data': {}
                    }
                    for i, column in enumerate(meta[1][1:]):
                        column = column.replace('.', '_')\
                            .replace('>=', 'p_')\
                            .replace('_ug/m3', '')\
                            .replace('_hpa', '')\
                            .replace('um/dl', '')\
                            .lower()
                        try:
                            value = int(line[i+1])
                        except ValueError:
                            try:
                                value = float(line[i+1])
                            except ValueError:
                                value = line[i+1]
                        entry['data'][column] = value
                    formatted['data'].append(entry)
                return formatted
            except:
                print('CSV format error!')
                return None
        else:
            print('Could not retrieve data from csv!')
            return None

    @staticmethod
    def load_csv(dat_file):
        try:
            with open(dat_file, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)
                meta, data = data[0:3], data[4:]
            return (meta, data)
        except:
            return None

    @staticmethod
    def get_epoch(time_val):
        # convert purple time elem to epoch timestamp
        pattern = '%Y-%m-%d %H:%M:%S'
        epoch = int(mktime(strptime(time_val, pattern)))
        return epoch


def write_txt(data):
    data_dir = 'data_history'
    t_file_num = 0
    entries = 0
    if not os.path.isdir(data_dir):
        os.mkdir(data_dir)
    else:
        shutil.rmtree(data_dir)
        os.mkdir(data_dir)

    history_file = open(f'{data_dir}/{data.get("data_name")}_{t_file_num}.txt', 'w')
    for entry in data.get('data'):
        if entries >= 30_000:
            history_file.close()
            print(f'{data.get("data_name")}_{t_file_num} written!')
            t_file_num += 1
            entries = 0
            history_file = open(f'{data_dir}/{data.get("data_name")}_{t_file_num}.txt', 'w')
            
        created_at = entry.get('created_at')
        for item in entry.get('data'):
            line = f'{data.get("host")} {item} {created_at} {entry["data"][item]}\n'
            history_file.write(line)
            entries += 1
    history_file.close()
    print(f'{data.get("data_name")}_{t_file_num} written!')
